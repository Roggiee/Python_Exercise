#第三题
from openpyxl import load_workbook
import os
def mmm(xlis):
    wb=load_workbook()
    ws=wb.worksheets[0]
    for a in ws.max_column:
        for n in ws.max_row:
            if ws.cell(row=n+1,column=c+1).value=="朱志豪":
                ws.cell(row=n+ 1, column=c+1).value ="朱智豪"
    wb.save(xlis)



#第四题
file_list=os.listdir(".")
for name in file_list:
    ws=name.rindex(".")
    if name[ws+1:]=="xlsx":
        modefyExcel(name)