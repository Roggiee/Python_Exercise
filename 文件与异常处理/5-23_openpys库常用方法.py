from openpyxl import Workbook
wb=Workbook()
ws=wb.worksheets[0]

#append()函数
row=["a","b","c","d"]
ws.append(row)

#读取Excel文件
from openpyxl import load_workbook
wc=load_workbook(r"e:\myexcel.xlsx")
ws=wb.worksheets[0]

#获取单元格内容
al=ws['AI'].value

#单元格赋值
ws['AI']=10
ws.cell(row=2,column=1,value=20)

#获取行和列
ws.max_row
ws.max_column

#遍历所有单元格
for r in range(ws.max_column):
    for c in range(ws.max_column):
        if c==ws.max_column:
            print("\n")
        print(ws.cell(row=r+1,column=c+1).value)